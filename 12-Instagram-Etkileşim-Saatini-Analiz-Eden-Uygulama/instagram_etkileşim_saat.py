"""
Instagram Etkileşim Saatini Analiz Eden Uygulama
=================================================
Bu uygulama, belirtilen bir Instagram hesabının gönderilerini çekerek
hangi gün, saat ve ayda paylaşım yapıldığını analiz eder.
Sonuçlar Excel dosyasına yazılır ve yüksek beğeni alan gönderiler
koşullu biçimlendirme ile yeşil renkte vurgulanır.

Kullanılan Kütüphaneler:
    - instaloader : Instagram profilinden gönderi verilerini çekmek için
    - pandas      : Verileri DataFrame yapısına dönüştürüp Excel'e yazmak için
    - openpyxl    : Excel dosyasını açıp koşullu biçimlendirme uygulamak için
    - getpass     : Şifreyi terminalde gizli (*) olarak almak için
    - locale      : Tarih bilgilerini Türkçe göstermek için
    - os          : Dosya yolu işlemleri için
"""

import os
import instaloader
import locale
import pandas as pd
from datetime import datetime
import getpass


# =============================================================================
# 1. ADIM: Türkçe Yerel Ayar (Locale) Yapılandırması
# =============================================================================
# strftime ile tarih formatlama yapıldığında gün ve ay isimlerinin
# Türkçe görünmesi için locale ayarını değiştiriyoruz.
# Linux/Mac: "tr_TR.UTF-8", Windows: "Turkish_Türkiye.1254"
try:
    locale.setlocale(locale.LC_ALL, "tr_TR.UTF-8")
except locale.Error:
    # Windows'ta Türkçe locale farklı isimle tanımlı olabilir
    try:
        locale.setlocale(locale.LC_ALL, "Turkish_Türkiye.1254")
    except locale.Error:
        print("Uyarı: Türkçe yerel ayar bulunamadı, varsayılan dil kullanılacak.")


# =============================================================================
# 2. ADIM: Instaloader Örneği Oluşturma ve Instagram'a Giriş Yapma
# =============================================================================
# Instagram anonim (giriş yapmadan) API isteklerini engellediği için
# önce kullanıcı adı ve şifre ile oturum açmak gerekiyor.
L = instaloader.Instaloader()

print("\n--- Instagram Giriş ---")
print("Not: Instagram anonim API isteklerini engellediği için giriş yapmanız gerekmektedir.")
ig_username = input("Instagram kullanıcı adınız: ")
# getpass.getpass() şifreyi terminalde gizli tutar (ekranda göstermez)
ig_password = getpass.getpass("Instagram şifreniz (gizli): ")

try:
    # Instaloader'ın login() metodu ile Instagram oturumu açılır
    L.login(ig_username, ig_password)
    print("Giriş başarılı!\n")
except instaloader.exceptions.BadCredentialsException:
    # Kullanıcı adı veya şifre hatalıysa
    print("Hata: Kullanıcı adı veya şifre yanlış!")
    exit(1)
except instaloader.exceptions.TwoFactorAuthRequiredException:
    # İki faktörlü doğrulama (2FA) aktifse
    print("Hata: İki faktörlü doğrulama gerekli. Instagram uygulamasından onay verin.")
    exit(1)
except Exception as e:
    # Diğer beklenmeyen hatalar
    print(f"Giriş hatası: {e}")
    exit(1)


# =============================================================================
# 3. ADIM: Analiz Edilecek Profil Bilgilerini Alma
# =============================================================================
# Kullanıcıdan analiz edilecek Instagram hesap adını ve gönderi sayısını alıyoruz.
username = input("Lütfen analiz etmek istediğiniz Instagram kullanıcı adını girin: ")
count = int(input("Lütfen analiz etmek istediğiniz gönderi sayısını girin: "))

# Profili yükle ve gönderilere erişim sağla
try:
    profile = instaloader.Profile.from_username(L.context, username)
    posts = profile.get_posts()
except instaloader.exceptions.ProfileNotExistsException:
    print(f"Hata: '{username}' adlı profil bulunamadı!")
    exit(1)
except Exception as e:
    print(f"Profil yükleme hatası: {e}")
    exit(1)


# =============================================================================
# 4. ADIM: Gönderi Verilerini Toplama ve Listeye Aktarma
# =============================================================================
# Her gönderiden gün, ay, yıl, saat, beğeni sayısı ve gönderi linki çekilir.
post_data = []

for i, post in enumerate(posts):
    # Belirlenen gönderi sayısına ulaşılınca döngüyü kır
    if i >= count:
        break

    # Her gönderi için sözlük (dict) yapısında veri oluştur
    post_info = {
        "Gün": post.date.strftime("%A"),           # Gönderinin paylaşıldığı gün adı (Pazartesi, Salı...)
        "Ay": post.date.strftime("%B"),             # Gönderinin paylaşıldığı ay adı (Ocak, Şubat...)
        "Yıl": post.date.strftime("%Y"),            # Gönderinin paylaşıldığı yıl (2024, 2025...)
        "Beğeni Sayısı": post.likes,                # Gönderinin toplam beğeni sayısı
        "Saat": post.date.strftime("%H:%M:%S"),     # Gönderinin paylaşıldığı saat (14:30:00 gibi)
        "Gönderi Linki": f'https://instagram.com/p/{post.shortcode}'  # Gönderinin doğrudan linki
    }
    post_data.append(post_info)
    print(f"Gönderi {i + 1} tamamlandı...")


# =============================================================================
# 5. ADIM: Pandas DataFrame Oluşturma ve Türkçe Karakter Düzeltmesi
# =============================================================================
# Toplanan verileri pandas DataFrame'e dönüştür
df = pd.DataFrame(post_data)

# Türkçe karakter sorununu ortadan kaldırma (encode/decode işlemi)
df['Gün'] = df['Gün'].str.encode('utf-8').str.decode('utf-8')
df['Ay'] = df['Ay'].str.encode('utf-8').str.decode('utf-8')


# =============================================================================
# 6. ADIM: Excel Dosyasına Veri Yazma
# =============================================================================
# Excel dosyasını script'in bulunduğu proje klasörüne kaydet
# (CWD'den bağımsız olarak her zaman doğru klasöre yazılır)
script_dir = os.path.dirname(os.path.abspath(__file__))
excel_file = os.path.join(script_dir, 'instagram_analist.xlsx')

# DataFrame'i Excel dosyasına yaz (openpyxl motoru ile)
df.to_excel(excel_file, index=False, engine='openpyxl')


# =============================================================================
# 7. ADIM: Excel'de Koşullu Biçimlendirme (Conditional Formatting)
# =============================================================================
# Beğeni sayısı 2000 ve üstü olan hücreleri yeşil renkle vurgulama
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Oluşturulan Excel dosyasını openpyxl ile tekrar aç
wb = load_workbook(excel_file)
ws = wb.active

# Yeşil renk dolgu tanımı (HEX: 00FF00)
yesil_dolgu = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# "Beğeni Sayısı" sütunu (4. sütun) üzerinde döngü ile koşullu biçimlendirme uygula
for row in ws.iter_rows(min_row=2, min_col=4, max_row=len(df) + 1, max_col=4):
    for cell in row:
        if cell.value >= 2000:
            cell.fill = yesil_dolgu

# Biçimlendirilmiş Excel dosyasını kaydet
wb.save(excel_file)


# =============================================================================
# 8. ADIM: Sonuç Bildirimi
# =============================================================================
print(f"\nVeriler {excel_file} dosyasına yazıldı!")
print("Beğeni sayısı 2000 ve üstü olanlar yeşil ile işaretlendi!")
input("\nİşlemi bitirmek için Enter tuşuna basın!")